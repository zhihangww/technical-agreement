"""
Excel 规范参数解析模块

将复杂结构的 Excel 规范文件解析为 HTML 表格文本，
通过大模型提取技术参数名称和规范值。

支持格式：.xls / .xlsx / .xlsm

功能：
  1. Excel 读取 → 合并单元格展平 → HTML 输出（ExcelParser）
  2. HTML 分块 → LLM 提取 → 参数名列表 + 规范库数据（ExcelParamExtractor）
"""

import json
import os
from pathlib import Path
from html import escape


class ExcelParser:
    """Excel 文件解析器，支持合并单元格展平、空行过滤、HTML 输出"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.ext = Path(file_path).suffix.lower()

        if self.ext not in ('.xls', '.xlsx', '.xlsm'):
            raise ValueError(f"不支持的文件格式: {self.ext}，仅支持 .xls / .xlsx / .xlsm")

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

    # ==============================================================
    # 公开接口
    # ==============================================================

    def get_sheet_names(self) -> list:
        """获取所有 Sheet 名称"""
        if self.ext == '.xls':
            return self._get_sheet_names_xls()
        else:
            return self._get_sheet_names_xlsx()

    def parse_sheet(self, sheet_name: str) -> str:
        """
        解析指定 Sheet，返回完整 HTML 表格字符串。
        处理流程：读取 → 展平合并单元格 → 裁剪空列 → 过滤空行 → 转 HTML
        """
        grid = self._read_sheet(sheet_name)
        grid = self._trim_trailing_empty_cols(grid)
        grid = self._filter_empty_rows(grid)
        return self._to_html(grid, sheet_name)

    def parse_sheet_to_chunks(self, sheet_name: str, rows_per_chunk: int = 100) -> list:
        """
        解析指定 Sheet，按行数分块返回多个 HTML 片段。
        用于后续分批喂给大模型处理。
        """
        grid = self._read_sheet(sheet_name)
        grid = self._trim_trailing_empty_cols(grid)
        grid = self._filter_empty_rows(grid)

        if not grid:
            return []

        chunks = []
        for i in range(0, len(grid), rows_per_chunk):
            chunk_rows = grid[i:i + rows_per_chunk]
            label = f"{sheet_name} (行 {i + 1}-{i + len(chunk_rows)})"
            chunks.append(self._to_html(chunk_rows, label))

        return chunks

    def get_sheet_stats(self, sheet_name: str) -> dict:
        """获取 Sheet 的统计信息"""
        grid = self._read_sheet(sheet_name)
        total_rows = len(grid)
        grid = self._trim_trailing_empty_cols(grid)
        grid = self._filter_empty_rows(grid)

        return {
            "sheet_name": sheet_name,
            "total_rows": total_rows,
            "non_empty_rows": len(grid),
            "columns": len(grid[0]) if grid else 0,
            "empty_rows_removed": total_rows - len(grid),
        }

    # ==============================================================
    # xlsx / xlsm 读取（openpyxl）
    # ==============================================================

    def _get_sheet_names_xlsx(self) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    def _read_xlsx(self, sheet_name: str) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb[sheet_name]

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        grid = []
        for r in range(1, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                row_data.append(str(val).strip() if val is not None else "")
            grid.append(row_data)

        for merged_range in ws.merged_cells.ranges:
            r1 = merged_range.min_row - 1
            r2 = merged_range.max_row
            c1 = merged_range.min_col - 1
            c2 = merged_range.max_col
            fill_value = grid[r1][c1]
            for r in range(r1, r2):
                for c in range(c1, c2):
                    grid[r][c] = fill_value

        wb.close()
        return grid

    # ==============================================================
    # xls 读取（xlrd）
    # ==============================================================

    def _get_sheet_names_xls(self) -> list:
        import xlrd
        wb = xlrd.open_workbook(self.file_path)
        return wb.sheet_names()

    def _read_xls(self, sheet_name: str) -> list:
        import xlrd

        try:
            wb = xlrd.open_workbook(self.file_path, formatting_info=True)
        except Exception:
            wb = xlrd.open_workbook(self.file_path)

        ws = wb.sheet_by_name(sheet_name)

        grid = []
        for r in range(ws.nrows):
            row_data = []
            for c in range(ws.ncols):
                cell_type = ws.cell_type(r, c)
                val = ws.cell_value(r, c)

                if val is None or (isinstance(val, str) and not val.strip()):
                    row_data.append("")
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    row_data.append(str(int(val)) if val == int(val) else str(val))
                else:
                    row_data.append(str(val).strip())
            grid.append(row_data)

        for (rlo, rhi, clo, chi) in ws.merged_cells:
            fill_value = grid[rlo][clo]
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    grid[r][c] = fill_value

        return grid

    # ==============================================================
    # 工具方法
    # ==============================================================

    def _read_sheet(self, sheet_name: str) -> list:
        """根据文件格式选择读取方式"""
        if self.ext == '.xls':
            return self._read_xls(sheet_name)
        else:
            return self._read_xlsx(sheet_name)

    def _filter_empty_rows(self, grid: list) -> list:
        """过滤全空行"""
        return [row for row in grid if any(cell for cell in row)]

    def _trim_trailing_empty_cols(self, grid: list) -> list:
        """裁剪右侧全空列，减少无效内容"""
        if not grid:
            return grid

        max_col_used = 0
        for row in grid:
            for c in range(len(row) - 1, -1, -1):
                if row[c]:
                    max_col_used = max(max_col_used, c)
                    break

        return [row[:max_col_used + 1] for row in grid]

    def _to_html(self, grid: list, title: str = "") -> str:
        """二维数组 → HTML table"""
        if not grid:
            return ""

        lines = []
        if title:
            lines.append(f"<h3>{escape(title)}</h3>")

        lines.append("<table border='1' cellpadding='4' cellspacing='0'>")
        for row in grid:
            lines.append("  <tr>")
            for cell in row:
                lines.append(f"    <td>{escape(cell) if cell else ''}</td>")
            lines.append("  </tr>")
        lines.append("</table>")

        return "\n".join(lines)


# ==================================================================
# LLM 提取 Prompt
# ==================================================================

EXCEL_EXTRACTION_PROMPT = """你是一个专业的电气设备技术参数文档分析专家。

## 任务
分析以下 HTML 表格，提取其中所有技术参数的中文名称、英文名称和规范值。

## 重要说明
- 表格结构可能不规整，行列可能有错开，**不要假设固定的列位置**
- 根据语义理解判断每个单元格的含义（中文名、英文名、数值、单位等）
- 设备类别（如"断路器"、"隔离开关"）可能出现在独立的列中，需要拼接为参数名前缀
  例如：设备列="断路器"，参数列="额定电压" → chinese_name="断路器额定电压"
- 英文参数名可能在相邻列、同一单元格的括号中、或独立的行中
- 规范值可能包含数字、单位、范围符号（≤、≥、~）等，需完整提取

## 文档内容
__DOCUMENT__

## 输出格式（严格 JSON）
{
    "parameters": [
        {
            "chinese_name": "断路器额定电压",
            "english_name": "Rated Voltage",
            "spec_value": "126kV"
        }
    ]
}

## 提取规则
1. 只提取实际的技术参数行，跳过标题行、说明文字、目录、备注
2. 如果某个参数没有英文名，english_name 填空字符串
3. 如果某个参数没有明确的规范值，spec_value 填空字符串
4. 设备类别需要拼接到中文参数名前面，形成完整唯一名称
5. 规范值要保留完整的数字+单位+符号，如 "≤28ms"、"3150A"、"4.8~5.8m/s"
6. 只输出 JSON，不要有其他内容"""


# ==================================================================
# LLM 参数提取器
# ==================================================================

class ExcelParamExtractor:
    """
    Excel 规范参数提取器

    流程：ExcelParser 解析 → HTML 分块 → LLM 逐块提取 → 合并去重
    输出：中文参数名列表、英文参数名列表、规范库条目
    """

    def __init__(self, model: str = "azure/gpt-4o"):
        self.model = model
        self.parser = None

    def load_file(self, file_path: str) -> list:
        """加载 Excel 文件，返回 Sheet 名称列表"""
        self.parser = ExcelParser(file_path)
        return self.parser.get_sheet_names()

    def extract(self, sheet_name: str, rows_per_chunk: int = 100,
                max_chunks: int = 0) -> dict:
        """
        提取指定 Sheet 中的参数。

        Args:
            sheet_name: 要处理的 Sheet 名称
            rows_per_chunk: 每个分块的行数
            max_chunks: 最多处理的块数，0 表示全部处理（用于调试省 token）

        Returns:
            {
                "chinese_names": ["参数1", "参数2", ...],
                "english_names": ["Param1", "Param2", ...],
                "spec_entries": [{"name": "...", "value": "...", "type": ""}, ...],
                "paired_names": [{"chinese": "...", "english": "..."}, ...],
                "total_extracted": int
            }
        """
        if not self.parser:
            raise RuntimeError("请先调用 load_file() 加载 Excel 文件")

        chunks = self.parser.parse_sheet_to_chunks(sheet_name, rows_per_chunk)
        if max_chunks > 0:
            chunks = chunks[:max_chunks]

        print(f"共 {len(chunks)} 个分块待处理")

        all_params = []
        for i, chunk in enumerate(chunks):
            print(f"  处理块 {i + 1}/{len(chunks)}...")
            extracted = self._extract_chunk(chunk)
            print(f"    提取到 {len(extracted)} 个参数")
            all_params.extend(extracted)

        unique_params = self._deduplicate(all_params)
        print(f"\n去重后共 {len(unique_params)} 个参数")

        return self._build_result(unique_params)

    # ==============================================================
    # 内部方法
    # ==============================================================

    def _extract_chunk(self, html_chunk: str) -> list:
        """对一个 HTML 块调用 LLM 提取参数"""
        from meri.utils.llm_utils import complete_chat

        prompt = EXCEL_EXTRACTION_PROMPT.replace("__DOCUMENT__", html_chunk)

        messages = [
            {
                "role": "system",
                "content": "你是专业的电气设备技术参数文档分析专家。请严格按要求输出JSON格式。"
            },
            {
                "role": "user",
                "content": [{"type": "text", "text": prompt}]
            }
        ]

        for attempt in range(3):
            try:
                response = complete_chat(
                    model=self.model,
                    messages=messages,
                    temperature=0.1,
                    response_format={"type": "json_object"},
                    max_tokens=8192
                )
                result = json.loads(response)
                return result.get("parameters", [])

            except json.JSONDecodeError:
                if attempt < 2:
                    print(f"    JSON 解析失败，重试 ({attempt + 2}/3)...")
                    continue
                print(f"    JSON 解析失败，跳过此块")
                return []

            except Exception as e:
                if attempt < 2:
                    print(f"    调用失败: {str(e)[:80]}，重试 ({attempt + 2}/3)...")
                    continue
                print(f"    调用失败，跳过此块: {str(e)[:80]}")
                return []

        return []

    def _deduplicate(self, params: list) -> list:
        """按中文参数名去重，保留首次出现的条目"""
        seen = set()
        unique = []
        for p in params:
            cn = (p.get("chinese_name") or "").strip()
            if cn and cn not in seen:
                seen.add(cn)
                unique.append(p)
        return unique

    def _build_result(self, params: list) -> dict:
        """将去重后的参数列表整理为最终输出格式"""
        chinese_names = []
        english_names = []
        spec_entries = []
        paired_names = []

        for p in params:
            cn = (p.get("chinese_name") or "").strip()
            en = (p.get("english_name") or "").strip()
            val = (p.get("spec_value") or "").strip()

            if cn:
                chinese_names.append(cn)
            if en:
                english_names.append(en)

            if cn:
                spec_entries.append({
                    "name": cn,
                    "value": val,
                    "type": ""
                })

            if cn or en:
                paired_names.append({
                    "chinese": cn,
                    "english": en
                })

        return {
            "chinese_names": chinese_names,
            "english_names": english_names,
            "spec_entries": spec_entries,
            "paired_names": paired_names,
            "total_extracted": len(params)
        }
